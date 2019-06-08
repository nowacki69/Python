import requests
import pandas as pd

from bs4 import BeautifulSoup
from openpyxl import load_workbook


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist 
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


def clear_spreadsheet(file):
    try:
        workbook = openpyxl.load_workbook(file)
        worksheets = workbook.sheetnames
        for worksheet in worksheets:
            if not worksheet == "Main":
                std = workbook[worksheet]
                workbook.remove(std)
                
        workbook.save('DVDsReleaseDates.xlsx')
        
    except:
        pass

filename = "DVDsReleaseDates.xlsx"
clear_spreadsheet(filename)

base_url = "https://www.dvdsreleasedates.com/top-movies-"

all_urls = []
first_url = base_url + "all-time/"

r = requests.get(first_url)
c = r.content
soup = BeautifulSoup(c, "html.parser")

all_movies = []
data = soup.find_all("td", {"class":"dvdcell"})
for item in data:
    movie = {}
    movie["Chart"] = "AllTime"
    movie["Position"] = int(list(item)[0])
    movie["Title"] = str(list(item)[4].string)
    movie["IMDB"] = float(item.find("td", {"class":"imdblink left"}).find("a").text)
    try:
        movie["Rating"] = item.find("td", {"class":"imdblink right"}).text.replace("\xa0", "")
    except:
        movie["Rating"] = None
    all_movies.append(movie)

dfAll = pd.DataFrame(all_movies, columns=["Position", "Title", "IMDB", "Rating"])

writer = pd.ExcelWriter(filename, mode="a")
dfAll.to_excel(writer, index=None, header=True, sheet_name='AllTime')
# append_df_to_excel(filename, dfAll, sheet_name="AllTime", index=False, header=True)

for i in range(2019, 1989, -1):
    url = base_url + str(i) + "/"
    r = requests.get(url)
    c = r.content
    soup = BeautifulSoup(c, "html.parser")

    all_movies = []
    data = soup.find_all("td", {"class":"dvdcell"})
    for item in data:
        movie = {}
        movie["Chart"] = str(i)
        movie["Position"] = int(list(item)[0])
        movie["Title"] = str(list(item)[4].string)

        try:
            movie["IMDB"] = float(item.find("td", {"class":"imdblink left"}).find("a").text)
        except:
            movie["IMDB"] = item.find("td", {"class":"imdblink left"}).find("a").text

        try:
            movie["Rating"] = item.find("td", {"class":"imdblink right"}).text.replace("\xa0", "")
        except:
            movie["Rating"] = None
        all_movies.append(movie)

    dfAll = pd.DataFrame(all_movies, columns=["Position", "Title", "IMDB", "Rating"])
    dfAll.to_excel(writer, index=None, header=True, sheet_name=str(i))
    # append_df_to_excel(filename, dfAll, sheet_name=str(i), index=False, header=True)

writer.save()
